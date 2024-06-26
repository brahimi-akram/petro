#the first part is a script that takes information from a sheet and create instances for each employe in that sheet If you want to use it you should modify model so that it won't check for pk
from django.db import connection
import win32com.client
import pythoncom
from .models import *
from openpyxl import load_workbook
from datetime import datetime 
import glob
import os
import subprocess
def init_sheet():
    Employe.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM sqlite_sequence WHERE name='pointage_employe';")

# Step 3: Vacuum the database to reclaim storage space
    with connection.cursor() as cursor:
        cursor.execute("VACUUM;")
    
    file=r'C:\Users\lenovo\OneDrive\Bureau\git_repo\pointage_for_windows\pointage\lISTE DU PERSONNEL.xlsx'
    workbook=load_workbook(file)
    sheet_names=workbook.sheetnames
    i=0
    column=['C','D','E','F']
    for sheet in sheet_names:
        worksheet=workbook[sheet]
        row=14
        i+=1
        while worksheet[f'{column[0]}{row}'].value:
            employe=Employe()
            employe.id=worksheet[f'{column[0]}{row}'].value
            employe.name=worksheet[f'{column[1]}{row}'].value
            employe.last_name=worksheet[f'{column[2]}{row}'].value
            employe.function=worksheet[f'{column[3]}{row}'].value
            employe.unite_id=i
            row+=1
            employe.save()

def init_code():
    Code_Employe.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM sqlite_sequence WHERE name='pointage_code_employe';")

# Step 3: Vacuum the database to reclaim storage space
    with connection.cursor() as cursor:
        cursor.execute("VACUUM;")
    current_dir = os.getcwd()
    parent_dir=os.path.join(current_dir,"test")
# Search for Excel files in the parent directory
    excel_files = glob.glob(os.path.join(parent_dir, '*.xlsx'))
    for file in excel_files:
        workbook=load_workbook(file)
        sheet_names = workbook.sheetnames

    # Print the list of sheet names
        for name in sheet_names:
            worksheet=workbook[name]
            try:
                emp=Employe.objects.get(pk=int(name))
            except:
                continue
            days=[]
            codes=[]
            date_current = worksheet['S4'].value
            year_of_the_file=date_current.year
            month_of_the_file=date_current.month
            try:
                if worksheet['B17'].value.isdigit():
                    index=17
                else:
                    index=18
            except:
                index=17

            for row in worksheet[f'B{index}:AF{index}']:
                for cell in row:
                    days.append(cell.value)
            for row in worksheet[f'B{index+1}:AF{index+1}']:
                for cell in row:
                    codes.append(cell.value)

            #if file=='C:\Users\lenovo\OneDrive\Bureau\pointage\test\Copie de FEUILLE DE POSITION ALGER JUIN 2024 (2)' and :
            for iterator in range(len(days)):
                
                code_emp=Code_Employe()
                code_iter=codes[iterator]
                try:
                    code_emp.code=Code.objects.get(pk=codes[iterator])
                except:
                    if codes[iterator] :
                        print(f'wrong code {codes[iterator]} in {file} for {emp.id} at {iterator}')
                    continue
                if int(iterator)>=16:
                    if date_current.month==12:
                        code_emp.date=date_current.replace(day=int(days[iterator]),month=1,year=year_of_the_file+1)
                    else:    
                        code_emp.date=date_current.replace(day=int(days[iterator]),month=date_current.month+1)
                else:
                    code_emp.date=date_current.replace(day=int(days[iterator]))
                code_emp.employe=emp
                code_emp.save()

def init_employe():
    current_dir = os.getcwd()
    parent_dir=os.path.join(current_dir,"creer")
# Search for Excel files in the parent directory
    excel_files = glob.glob(os.path.join(parent_dir, '*.xlsx'))
    if excel_files:
        for file in excel_files:
            workbook=load_workbook(file)
            sheet_names=workbook.sheetnames
            for sheet_name in sheet_names:
                worksheet=workbook[sheet_name]
                if worksheet['B7'].value:
                    try:
                        employe=Employe.objects.get(name=worksheet['B7'].value ,last_name=worksheet['F7'].value)
                    except:
                        employe=Employe()
                    employe.name=worksheet['B7'].value
                    employe.last_name=worksheet['F7'].value
                    employe.date_of_birth=worksheet['B9'].value
                    employe.place_of_birth=worksheet['D9'].value
                    employe.wilaya_of_birth=worksheet['H9'].value
                    employe.adresse=worksheet['B11'].value
                    employe.adresse_wilaya=worksheet['H11'].value
                    employe.father_name=worksheet['B13'].value
                    employe.mother_name=worksheet['E13'].value
                    employe.phone=worksheet['I13'].value
                    employe.familiy_situation=worksheet['B15'].value
                    employe.numbre_of_children=worksheet['D15'].value
                    employe.blood_type=worksheet['F15'].value
                    employe.cnas_number=worksheet['H15'].value
                    employe.function=worksheet['B17'].value
                    employe.position=worksheet['G17'].value
                    if not employe.unite:
                        unite=["ALGER","MECHRIA","TAMANRASSET","AIN SEFRA","IN SALAH"]
                        index=unite.index(employe.position)
                        employe.unite=Unite.objects.get(id=index+1)

                    employe.enterprise=worksheet['I17'].value
                    employe.recruitment_date=worksheet['B19'].value
                    employe.department=worksheet['E19'].value
                    employe.service=worksheet['H19'].value
                    employe.contract_number=worksheet['B21'].value
                    employe.contract_effective_date=worksheet['D21'].value
                    employe.contract_validation_date=worksheet['F21'].value
                    employe.contract_termination_date=worksheet['H21'].value
                    if worksheet['C23'].value != 'EXEMPLTE':
                        employe.national_service_departure_date=worksheet['C23'].value
                        employe.national_service_returne_date=worksheet['E23'].value
                        employe.national_service_recall_departure_date=worksheet['G23'].value
                        employe.national_service_recallt_return_date=worksheet['I23'].value
                    employe.account_number=worksheet['D25'].value
                    employe.account_key=worksheet['G25'].value
                    employe.account_agency=worksheet['I25'].value
                    if worksheet['C27'].value:
                        employe.driver_license_number=worksheet['C27'].value
                        employe.driver_license_established_date=worksheet['E27'].value
                        employe.driver_license_experation_date=date_handling(worksheet['G27'].value)
                        employe.driver_license_type=worksheet['I27'].value
                    employe.cni_number=worksheet['B29'].value
                    employe.cni_established_date=worksheet['E29'].value
                    employe.cni_established_by=worksheet['G29'].value
                    employe.save()
                    cell=worksheet['A33'].value
                    row=33
                    while cell:
                        diplome=Diplome()
                        diplome.establishment=cell
                        if worksheet[f'D{row}'].value:
                            diplome.entry_date=worksheet[f'D{row}'].value
                        diplome.end_date=date_handling(worksheet[f'E{row}'].value)
                        diplome.diplome_name=worksheet[f'F{row}'].value
                        diplome.id_employe=employe
                        diplome.save()
                        row+=1
                        cell=worksheet[f'A{row}'].value
                    if worksheet['B42'].value:
                        partner=Partner()
                        partner.id_employe=employe
                        partner.name=worksheet['B42'].value
                        partner.last_name=worksheet['F42'].value
                        partner.date_of_birth=worksheet['B44'].value
                        partner.place_of_birth=worksheet['F44'].value
                        partner.wilaya_of_birth=worksheet['I44'].value
                        partner.marriage_date=date_handling(worksheet['C46'].value)
                        partner.partner_salary=worksheet['F46'].value
                        partner.save()
                    cell=worksheet['A51'].value
                    row=51
                    while cell:
                        child=Child()
                        child.id_employe=employe
                        child.name=worksheet[f'A{row}'].value
                        child.last_name=worksheet[f'B{row}'].value
                        
                        child.date_of_birth=worksheet[f'D{row}'].value
                        student=worksheet['F51'].value
                        if student:
                            if student == 'NON':
                                child.student=False
                            else:
                                child.student=True
                        child.af=worksheet[f'I{row}'].value
                        child.save()
                        row+=1
                        cell=worksheet[f'A{row}'].value

def excel_to_pdf(pdf_path,output_file,output_folder):
    pythoncom.CoInitialize()
    excel_file=os.path.join(output_folder,output_file)
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    # Open the XLSX file
    workbook = excel.Workbooks.Open(excel_file)

    try:
        # Save the file as PDF
        workbook.ExportAsFixedFormat(0, pdf_path)
    finally:
        # Close the workbook and quit Excel
        workbook.Close(False)
        excel.Quit()

    '''try:
        # Construct the command to execute
        command = [file_path] + list(args)
        
        # Execute the command
        process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = process.communicate()
        # Check if the process exited successfully
        if process.returncode == 0:
            # Process output
            output = stdout.decode('utf-8')
            return output
        else:
            # Error occurred
            error = stderr.decode('utf-8')
            print(error)
            return f"Error: {error}"
    except Exception as e:
        print(e)
        return f"Error: {str(e)}"'''
    
def date_handling(input):
    if isinstance(input, datetime):
        output= input.date()  # Convert to date
    elif input:
        try:
                # Assuming date is in 'MM/DD/YYYY' format
            output = datetime.strptime(str(input), '%m/%d/%Y').date()
        except ValueError:
            output= None  # Handle invalid date formats
    return output


def codage(table,info,operation):
    current_directory=os.getcwd() #
    filename='changes.txt'
    file_path=os.path.join(current_directory,filename)
    try:
        with open (file_path,'a') as file3:
            string = str(table)+'||'+str(operation)+'||'
            file3.write (string)
            if table == 68:#employe
                field_names = [field.name for field in Employe._meta.get_fields() if not field.many_to_many and not field.one_to_many ]
            else :
                field_names = [field.name for field in Code_Employe._meta.get_fields() if not field.many_to_many and not field.one_to_many]
            
            data={field : getattr(info,field) for field in field_names}
            for key in data.keys():
                file3.write (str(data[key])+'||')
                
            file3.write (str('\n'))
        return 1
    except :
        return 0
    
def decodage():

    current_directory=os.getcwd() #
    filename='changes.txt'
    file_path=os.path.join(current_directory,filename)
    try:
        
        with open(file_path,'r') as file3:
            for line in file3:
                list_elements=line.split('||')
                if  list_elements[0]==str(68):#employe
                    field_names = [field.name for field in Employe._meta.get_fields() if not field.many_to_many and not field.one_to_many ]
                else :
                    field_names = [field.name for field in Code_Employe._meta.get_fields() if not field.many_to_many and not field.one_to_many]
                del list_elements[0]
                operation=list_elements.pop(0)
                if operation ==str(1):
                    code_emp=Code_Employe()
    except:
        pass
    return 1

def create_log(profile,operation,table):
    new_log=History()
    new_log.user=profile
    new_log.operation=operation
    new_log.table=table
    new_log.date=datetime.now()
    new_log.save()
        